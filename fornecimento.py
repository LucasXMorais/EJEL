from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import openpyxl

agrupamento = input('O projeto tem agrupamento coletivo?\n(s/n) : ')

planilha = "Fornecimento.xlsx"

if agrupamento == 's':
    planilha = 'Fornecimento_C_Agrupamento.xlsx'


#Cria um 'Workbook' usado pelo openpyxl
wb = Workbook()
#Carrega a planilha
wb = load_workbook(planilha,data_only = True)
#cria uma planilha ativa - 'Worksheet'
ws = wb.active

fornecimento = []
row = 0 

#Varre a planilha e cria uma lista com uma lista para cada circuito
for cell in ws['A']:    
    row += 1
    
    if cell.value == "A" or cell.value == "B" or cell.value == "C":        
        fornecimento_n =[]
        
        for column in ws.iter_cols(min_col=1, max_col=10, min_row=row, max_row=row, values_only=True):
            fornecimento_n.append(column[0])
            
        fornecimento.append(fornecimento_n)    

print(fornecimento) 


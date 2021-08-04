from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
#import numpy as np

def is_number(entrada):

    try:
        complex(entrada)

    except:

        return False

    return True


# Cria um 'Workbook' usado pelo openpyxl
wb = Workbook()

# Carrega a planilha
wb = load_workbook('Planilha teste.xlsx', data_only=True)

# cria uma planilha ativa - 'Worksheet'
#ws = wb.active #Trabalha com a planilha que está aberta
ws = wb['Planilha CAD'] #Abre a planilha específica ----- O nome deve ser exatamente igual de ambas as planilhas

row = 1 # Linha que o programa começa

circuitos = []

# Varre a planilha e cria uma lista com uma lista para cada circuito
for cell in ws.iter_rows(min_col=1, max_col=1, min_row=row, values_only=True): 
    
    if is_number(cell[0]):

        circuito_n = []
        
        for column in ws.iter_cols(min_col=1, max_col=14, min_row=row, max_row=row, values_only=True):                    

            if column[0] == None:

                circuito_n.append(0)

            else:

                circuito_n.append(column[0])
               
        if circuito_n.count(0) < 13:        
            circuitos.append(circuito_n)            

    else:
        if circuitos != []:
    
            break
            
    row += 1

print(circuitos)

''' # Varre a planilha e cria uma lista com uma lista para cada circuito
for cell in ws.iter_rows(min_col=1, max_col=1, min_row=row, values_only=True): 
    
    if is_number(cell[0]):

        circuito_n = []
        
        for column in ws.iter_cols(min_col=1, max_col=14, min_row=row, max_row=row, values_only=True):                    

            if column[0] == None:

                circuito_n.append(0)

            else:

                circuito_n.append(column[0])
               
        if circuito_n.count(0) < 13:        
            circuitos.append(circuito_n)            

    else:
        if circuitos != []:
    
            break
            
    row += 1

print(circuitos) '''
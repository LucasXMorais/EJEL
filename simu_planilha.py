from ler_planilha import ler_qdc
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl

#
#!!!!!!!!LEMBRE DE MUDAR O NOME DO ARQUIVO 'ler_planilha_(data).py' para 'ler_planilha.py'!!!!!!!!!!!
#

# Cria um 'Workbook' usado pelo openpyxl
wb = Workbook()
#
# Carrega a planilha
wb = load_workbook('Planilha teste.xlsx', data_only=True)
#
# cria uma planilha ativa - 'Worksheet'
#ws = wb.active #Trabalha com a planilha que está aberta
ws = wb['Planilha CAD'] #Abre a planilha específica ----- O nome deve ser exatamente igual de ambas as planilhas

#row = int(input('Qual a linha de início? '))
row = 1
i = 1

while 1:
     
     linha, qdc = ler_qdc(row, ws)
 
     if qdc == []:
          break
     
     print("QDC "+str(i))
     print(qdc) 
     
     i += 1    
     row = linha


''' 
You can do this in 2 ways. First is just to import the specific function you want from file.py. To do this use

     from file import function

Another way is to import the entire file

     import file as fl
Then you can call any function inside file.py using

     fl.function(a,b) 
'''

#Esse comentário, é um comentário que achei explicando como importar as funções
#Basciamente podemos usar dois modos. Acho que o primeiro seria bom pra essa função, pois o arquivo só tem uma
#E o segundo seria melhor para algum arquivo com mts funções
#Nesse arquivo usei o primeiro método
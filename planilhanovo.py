from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl


def is_number(entrada):
    
    try:
        complex(entrada)
    except:
        return False
    
    return True

#Cria um 'Workbook' usado pelo openpyxl
wb = Workbook()
#Carrega a planilha
wb = load_workbook('Planilha teste.xlsx',data_only = True)
#cria uma planilha ativa - 'Worksheet'
ws = wb['Planilha CAD']
#ws = wb.active

circuitos = []
row = 0 

#Varre a planilha e cria uma lista com uma lista para cada circuito
for cell in ws['A']:    
    row += 1
    
    if is_number(cell.value):        
        circuito_n =[]
        
        for column in ws.iter_cols(min_col=1, max_col=14, min_row=row, max_row=row, values_only=True):
            if column[0] == None:
                circuito_n.append(0)
            else:
                circuito_n.append(column[0])
            
        circuitos.append(circuito_n)    

print(circuitos) 


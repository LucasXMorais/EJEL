from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('sheet1')
ws2 = wb.create_sheet('sheet2')
ws3 = wb.create_sheet('sheet3')

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row = x, column = y)

columns = list(range(0,10))

for col in ws.iter_cols(min_col = 4, max_col = 6, max_row = 2):
    for cell in col:
        ws1.append(columns)
        

wb.save('xl.xlsx')